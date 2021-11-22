import re

import openpyxl

wb = openpyxl.reader.excel.load_workbook(filename='test.xlsx', data_only=True)
count_of_lists = len(wb.sheetnames)  # количество листов в книге


def input_lists(msg):
    input_data = input(msg)
    if not input_data.isdigit(): return input_lists("Вы ввели не число. Введите целое положительное число: ")
    if 0 <= int(input_data) <= count_of_lists: return int(input_data) - 1
    return input_lists(f'Всего листов в книге: {count_of_lists}.'
                       f' Введите число от 1 до {count_of_lists}. ')


def proj_validation_check(proj_list):
    """Функция делает проверку на количество символов в ID проекта.Удаляет спец.символы и пробелы"""
    correct_id = []
    for id_proj in proj_list:
        val_id = re.sub(r'[^\w]', '', str(id_proj))
        correct_id.append(val_id.strip())
    return correct_id


def win_rate():
    projects = []
    projects_win = []
    list_excel = input_lists("Введите номер листа Ecxel, в котором данные о проектах: ")
    sheet = wb.worksheets[list_excel]  # Обращение к определенному листу. 0 - это 1-й лист
    for row in range(2, sheet.max_row + 1):  # max_row - проходит до конца значений по столбцу
        id_project = sheet[row][0].value  # 0,1,2...n - идентификаторы столбцов слева направо
        if id_project is not None:
            projects.append(id_project)
        id_project_win = sheet[row][1].value
        if id_project_win is not None:
            projects_win.append(id_project_win)
    projects = proj_validation_check(set(projects))  # удаляем повторы в списке и спецсимволы
    projects_win = proj_validation_check(set(projects_win))  # удаляем повторы в списке и спецсимволы
    count_projects = len(projects)
    count_projects_win = len(projects_win)
    if count_projects != 0:
        projects_win_rate = round(count_projects_win / count_projects * 100, 1)  # round - округление до одного знака
        print(f'Зарегистрированных проектов: {count_projects}')
        print(f'Зарегистрированные проекты: {projects}')
        print(f'Выигранных проектов: {count_projects_win}')
        print(f'Выигранные проекты: {projects_win}')
        print(f'WinRate равен: {projects_win_rate} %')
    else:
        projects_win_rate = 0
        return print(f'Количество зарегистрированных проектов равно {projects_win_rate}. Проверьте файл!')


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    win_rate()
